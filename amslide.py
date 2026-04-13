from __future__ import annotations

import io
import json
import re
from copy import copy
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from simple_salesforce import Salesforce

TERM_SHEET = "Term"
BRIDGE_SHEET = "Bridge"
DEFAULT_TEMPLATE_NAME = "Reference AM Templates.xlsx"
VALID_STAGES = ["Closed Won", "Expired", "Matured", "Paid Off", "Sold"]

TERM_RT_EXACT = {"term loan", "dscr"}
TERM_RT_CONTAINS = {"dscr"}
BRIDGE_RT_EXACT = {
    "acquired bridge loan",
    "bridge loan",
    "sab loan",
    "single asset bridge loan",
}
BRIDGE_RT_CONTAINS = {"sab", "single asset bridge"}
BRIDGE_DEV_EXACT = {"single_asset_bridge_loan"}
BRIDGE_DEV_CONTAINS = {"single_asset_bridge", "sab"}

DC_DEAL_FIELD = "Deal__c"


# -------------------------
# Basic helpers
# -------------------------
def install_truststore() -> None:
    try:
        import truststore

        truststore.inject_into_ssl()
    except Exception:
        pass



def soql_quote(value: str) -> str:
    return "'" + str(value).replace("\\", "\\\\").replace("'", "\\'") + "'"



def digits_only(value: Any) -> str:
    return re.sub(r"\D", "", "" if value is None or pd.isna(value) else str(value))



def last5_strip_prefix(value: Any) -> str:
    digits = digits_only(value)
    if digits.startswith("4030") or digits.startswith("6000"):
        digits = digits[4:]
    return digits[-5:] if len(digits) >= 5 else digits



def pct_to_dec(value: Any) -> float | None:
    if value in ("", None) or pd.isna(value):
        return None
    text = str(value).strip().replace("%", "")
    try:
        number = float(text)
        return number / 100.0 if number > 1.5 else number
    except Exception:
        return None



def parse_date_any(value: Any):
    if value in ("", None) or pd.isna(value):
        return None
    dt = pd.to_datetime(value, errors="coerce")
    if pd.isna(dt):
        return None
    return dt.date()



def extract_states_only(state_percentages: Any) -> str:
    text = "" if state_percentages is None or pd.isna(state_percentages) else str(state_percentages)
    codes = re.findall(r"\b[A-Z]{2}\b", text.upper())
    if not codes:
        return text.strip()
    seen: set[str] = set()
    out: list[str] = []
    for code in codes:
        if code not in seen:
            seen.add(code)
            out.append(code)
    return ", ".join(out)



def norm_rt(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)) or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())



def safe_flatten_recordtype(df: pd.DataFrame) -> pd.DataFrame:
    if "RecordType" in df.columns:
        df["RecordType.Name"] = df["RecordType"].apply(lambda x: (x or {}).get("Name"))
        df["RecordType.DeveloperName"] = df["RecordType"].apply(
            lambda x: (x or {}).get("DeveloperName")
        )
        df = df.drop(columns=["RecordType"], errors="ignore")
    return df



def chunked(values, size: int = 200):
    values = list(values)
    for i in range(0, len(values), size):
        yield values[i : i + size]



def build_where_for_search(mode: str, query_text: str) -> str:
    query_text = (query_text or "").strip()
    if mode == "Account Name":
        return "Account_Name__c LIKE " + soql_quote("%" + query_text + "%")
    if mode == "Deal Name":
        return "Name LIKE " + soql_quote("%" + query_text + "%")

    digits = re.sub(r"\D", "", query_text)
    if digits:
        return "(" + " OR ".join(
            [
                "Deal_Loan_Number__c = " + soql_quote(digits),
                "Deal_Loan_Number__c LIKE " + soql_quote("%" + digits + "%"),
                "Deal_Loan_Number__c LIKE " + soql_quote("%" + query_text + "%"),
            ]
        ) + ")"
    return "Deal_Loan_Number__c LIKE " + soql_quote("%" + query_text + "%")



def try_query_drop_missing(
    sf: Salesforce,
    object_name: str,
    fields: list[str],
    where_clause: str,
    limit: int = 2000,
    order_by: str | None = None,
):
    fields = list(fields)
    while True:
        soql = f"SELECT {', '.join(fields)} FROM {object_name} WHERE {where_clause}"
        if order_by:
            soql += f" ORDER BY {order_by}"
        soql += f" LIMIT {int(limit)}"
        try:
            rows = sf.query_all(soql).get("records", [])
            return rows, fields, soql
        except Exception as exc:
            message = str(exc)
            missing_column = re.search(r"No such column '([^']+)'", message)
            bad_relationship = re.search(r"Didn't understand relationship '([^']+)'", message)

            if missing_column:
                bad = missing_column.group(1)
                if bad in fields:
                    fields.remove(bad)
                    continue

            if bad_relationship:
                relbad = bad_relationship.group(1)
                to_drop = [
                    field
                    for field in fields
                    if field.startswith(relbad + ".") or (("." + relbad + ".") in field)
                ]
                if to_drop:
                    for field in to_drop:
                        fields.remove(field)
                    continue
            raise



def classify_term_bridge(df_all: pd.DataFrame):
    if "RecordType.Name" not in df_all.columns:
        return df_all.iloc[0:0].copy(), df_all.iloc[0:0].copy()

    rt_name = df_all["RecordType.Name"].apply(norm_rt)
    rt_dev = df_all.get(
        "RecordType.DeveloperName",
        pd.Series([""] * len(df_all), index=df_all.index),
    )
    rt_dev = rt_dev.fillna("").astype(str).str.strip().str.lower()

    term_exact = rt_name.isin(TERM_RT_EXACT)
    term_contains = pd.Series(False, index=df_all.index)
    for token in TERM_RT_CONTAINS:
        term_contains = term_contains | rt_name.str.contains(re.escape(token), na=False)
    is_term = term_exact | term_contains

    bridge_exact = rt_name.isin(BRIDGE_RT_EXACT)
    bridge_contains = pd.Series(False, index=df_all.index)
    for token in BRIDGE_RT_CONTAINS:
        bridge_contains = bridge_contains | rt_name.str.contains(re.escape(token), na=False)

    bridge_dev_exact = rt_dev.isin(BRIDGE_DEV_EXACT)
    bridge_dev_contains = pd.Series(False, index=df_all.index)
    for token in BRIDGE_DEV_CONTAINS:
        bridge_dev_contains = bridge_dev_contains | rt_dev.str.contains(re.escape(token), na=False)

    is_bridge = bridge_exact | bridge_contains | bridge_dev_exact | bridge_dev_contains
    is_bridge = is_bridge & (~is_term)
    return df_all[is_term].copy(), df_all[is_bridge].copy()


# -------------------------
# Salesforce connection
# -------------------------
def load_salesforce_oauth_config() -> dict[str, str]:
    try:
        secrets_section = dict(st.secrets.get("salesforce", {}))
    except Exception:
        secrets_section = {}

    required_keys = ["client_id", "client_secret", "auth_host", "redirect_uri"]
    missing_keys = [key for key in required_keys if not secrets_section.get(key)]
    if missing_keys:
        raise RuntimeError(
            "Missing Salesforce OAuth secrets: " + ", ".join(missing_keys)
            + ". Add them under [salesforce] in Streamlit secrets."
        )
    return secrets_section



def build_salesforce_login_url(oauth_config: dict[str, str]) -> str:
    auth_host = str(oauth_config["auth_host"]).rstrip("/")
    query = urlencode(
        {
            "response_type": "code",
            "client_id": oauth_config["client_id"],
            "redirect_uri": oauth_config["redirect_uri"],
            "scope": oauth_config.get("scope", "api refresh_token"),
            "prompt": oauth_config.get("prompt", "login"),
        }
    )
    return f"{auth_host}/services/oauth2/authorize?{query}"



def exchange_salesforce_code_for_token(
    oauth_config: dict[str, str],
    code: str,
) -> dict[str, Any]:
    install_truststore()
    auth_host = str(oauth_config["auth_host"]).rstrip("/")
    token_url = f"{auth_host}/services/oauth2/token"
    payload = urlencode(
        {
            "grant_type": "authorization_code",
            "client_id": oauth_config["client_id"],
            "client_secret": oauth_config["client_secret"],
            "redirect_uri": oauth_config["redirect_uri"],
            "code": code,
        }
    ).encode("utf-8")
    request = Request(
        token_url,
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        detail = body
        try:
            parsed = json.loads(body)
            detail = parsed.get("error_description") or parsed.get("error") or body
        except Exception:
            pass
        raise RuntimeError(f"Salesforce login failed: {detail}") from exc



def read_query_param(name: str) -> str | None:
    value = st.query_params.get(name)
    if isinstance(value, list):
        return value[0] if value else None
    return value



def clear_query_params() -> None:
    try:
        st.query_params.clear()
    except Exception:
        pass



def clear_salesforce_session() -> None:
    for key in [
        "salesforce_auth",
        "account_candidates",
        "term_preview",
        "bridge_preview",
        "occupancy_debug",
        "occupancy_period_summary",
        "period_labels",
        "workbook_bytes",
        "workbook_name",
        "match_count",
        "term_count",
    ]:
        st.session_state.pop(key, None)



def maybe_finish_salesforce_oauth(oauth_config: dict[str, str]) -> None:
    oauth_error = read_query_param("error")
    if oauth_error:
        oauth_description = read_query_param("error_description") or oauth_error
        clear_query_params()
        raise RuntimeError(f"Salesforce login was not completed: {oauth_description}")

    code = read_query_param("code")
    if not code:
        return

    if st.session_state.get("_last_salesforce_code") == code and st.session_state.get("salesforce_auth"):
        clear_query_params()
        return

    token_payload = exchange_salesforce_code_for_token(oauth_config, code)
    access_token = token_payload.get("access_token")
    instance_url = token_payload.get("instance_url")
    if not access_token or not instance_url:
        raise RuntimeError("Salesforce login succeeded, but no access token or instance URL was returned.")

    st.session_state["salesforce_auth"] = {
        "access_token": access_token,
        "instance_url": instance_url,
        "issued_at": token_payload.get("issued_at"),
        "id_url": token_payload.get("id"),
        "signature": token_payload.get("signature"),
    }
    st.session_state["_last_salesforce_code"] = code
    clear_query_params()
    st.rerun()



def get_salesforce_client_from_session() -> Salesforce | None:
    install_truststore()
    auth_data = st.session_state.get("salesforce_auth", {})
    instance_url = auth_data.get("instance_url")
    access_token = auth_data.get("access_token")
    if not instance_url or not access_token:
        return None
    return Salesforce(instance_url=instance_url, session_id=access_token)



def normalize_salesforce_error(exc: Exception) -> RuntimeError:
    message = str(exc)
    if "INVALID_SESSION_ID" in message or "Session expired" in message:
        clear_salesforce_session()
        return RuntimeError("Your Salesforce session expired. Click 'Log in to Salesforce' and try again.")
    return RuntimeError(message)


# -------------------------
# Query helpers
# -------------------------
def search_matching_accounts(sf: Salesforce, search_mode: str, query_text: str) -> pd.DataFrame:
    preview_fields = [
        "Id",
        "Name",
        "Deal_Loan_Number__c",
        "Account_Name__c",
        "RecordType.Name",
        "RecordType.DeveloperName",
        "StageName",
        "CloseDate",
    ]
    where_search = (
        build_where_for_search(search_mode, query_text)
        + " AND StageName IN ("
        + ", ".join(soql_quote(stage) for stage in VALID_STAGES)
        + ")"
    )
    preview_rows, _, _ = try_query_drop_missing(
        sf,
        "Opportunity",
        preview_fields,
        where_search,
        limit=2000,
        order_by="CloseDate DESC NULLS LAST",
    )
    if not preview_rows:
        return pd.DataFrame(columns=["Account_Name__c", "loans"])

    df_preview = pd.DataFrame(preview_rows).drop(columns=["attributes"], errors="ignore")
    df_preview = safe_flatten_recordtype(df_preview)
    df_preview = df_preview[df_preview["Account_Name__c"].notna()].copy()
    account_counts = (
        df_preview.groupby("Account_Name__c", dropna=False)
        .size()
        .reset_index(name="loans")
        .sort_values(["loans", "Account_Name__c"], ascending=[False, True])
        .reset_index(drop=True)
    )
    return account_counts



def find_contact_ref_field_on_deal_contact(sf: Salesforce):
    description = sf.Deal_Contact__c.describe()
    for field in description.get("fields", []):
        if field.get("type") == "reference":
            reference_to = field.get("referenceTo") or []
            if any(str(item).lower() == "contact" for item in reference_to):
                return field.get("name"), field.get("relationshipName")
    return None, None



def query_deal_contacts_for_guarantors(sf: Salesforce, opportunity_ids: list[str]) -> pd.DataFrame:
    if not opportunity_ids:
        return pd.DataFrame(columns=[DC_DEAL_FIELD, "GuarantorName"])

    _, contact_relationship_name = find_contact_ref_field_on_deal_contact(sf)
    contact_name_path = (
        f"{contact_relationship_name}.Name" if contact_relationship_name else None
    )

    fields = ["Id", DC_DEAL_FIELD, "Is_Guarantor__c", "Name"]
    if contact_name_path:
        fields.append(contact_name_path)

    rows: list[dict[str, Any]] = []
    for group in chunked(opportunity_ids, 150):
        ids_in = ", ".join(soql_quote(item) for item in group)
        trial_fields = list(fields)
        while True:
            soql = (
                "SELECT "
                + ", ".join(trial_fields)
                + " FROM Deal_Contact__c"
                + f" WHERE {DC_DEAL_FIELD} IN ({ids_in})"
                + " AND Is_Guarantor__c = TRUE"
            )
            try:
                result = sf.query_all(soql)
                rows.extend(result.get("records", []))
                break
            except Exception as exc:
                message = str(exc)
                missing_column = re.search(r"No such column '([^']+)'", message)
                bad_relationship = re.search(r"Didn't understand relationship '([^']+)'", message)

                if missing_column and missing_column.group(1) in trial_fields:
                    trial_fields.remove(missing_column.group(1))
                    continue

                if bad_relationship:
                    relbad = bad_relationship.group(1)
                    to_drop = [
                        field
                        for field in trial_fields
                        if field.startswith(relbad + ".") or (("." + relbad + ".") in field)
                    ]
                    if to_drop:
                        for field in to_drop:
                            trial_fields.remove(field)
                        continue
                raise

    df = pd.DataFrame(rows).drop(columns=["attributes"], errors="ignore")
    if df.empty:
        return pd.DataFrame(columns=[DC_DEAL_FIELD, "GuarantorName"])

    if contact_relationship_name and contact_relationship_name in df.columns:
        df["ContactName"] = df[contact_relationship_name].apply(lambda x: (x or {}).get("Name"))
    else:
        df["ContactName"] = None

    df["GuarantorName"] = df["ContactName"]
    missing_mask = df["GuarantorName"].isna() | (df["GuarantorName"].astype(str).str.strip() == "")
    df.loc[missing_mask, "GuarantorName"] = df.loc[missing_mask, "Name"]
    return df[[DC_DEAL_FIELD, "GuarantorName"]].copy()



def build_term_bridge_for_account(sf: Salesforce, account_name: str):
    opportunity_fields = [
        "Id",
        "Name",
        "Deal_Loan_Number__c",
        "Account_Name__c",
        "RecordType.Name",
        "RecordType.DeveloperName",
        "StageName",
        "CloseDate",
        "Amount",
        "Current_UPB__c",
        "UW_LTV__c",
        "Rate__c",
        "Current_Svc_Interest_Rate__c",
        "State_Percentages__c",
        "Total_Properties__c",
        "Total_Units__c",
        "Recourse__c",
        "Historical_Ontime_Payments_Percentage__c",
        "Next_Payment_Date__c",
        "Stated_Maturity_Date__c",
        "Original_Line_Maturity_Date__c",
        "Aggregate_Funding__c",
    ]

    where_account = (
        "Account_Name__c = "
        + soql_quote(account_name)
        + " AND StageName IN ("
        + ", ".join(soql_quote(stage) for stage in VALID_STAGES)
        + ")"
    )

    rows, _, _ = try_query_drop_missing(
        sf,
        "Opportunity",
        opportunity_fields,
        where_account,
        limit=2000,
        order_by="CloseDate DESC NULLS LAST",
    )

    df_all = pd.DataFrame(rows).drop(columns=["attributes"], errors="ignore")
    df_all = safe_flatten_recordtype(df_all)
    if df_all.empty:
        return pd.DataFrame(), pd.DataFrame()

    df_all["InterestRate_Picked"] = df_all.get("Rate__c")
    blank_rate = df_all["InterestRate_Picked"].isna() | (
        df_all["InterestRate_Picked"].astype(str).str.strip() == ""
    )
    df_all.loc[blank_rate, "InterestRate_Picked"] = df_all.loc[
        blank_rate, "Current_Svc_Interest_Rate__c"
    ]

    df_all["LTV_Dec"] = df_all.get("UW_LTV__c").apply(pct_to_dec) if "UW_LTV__c" in df_all.columns else None
    df_all["Rate_Dec"] = df_all["InterestRate_Picked"].apply(pct_to_dec)
    df_all["OriginationDate_dt"] = df_all.get("CloseDate").apply(parse_date_any)
    df_all["NextPay_dt"] = df_all.get("Next_Payment_Date__c").apply(parse_date_any)
    df_all["Maturity_Picked_raw"] = df_all.get("Stated_Maturity_Date__c")
    missing_maturity = df_all["Maturity_Picked_raw"].isna() | (
        df_all["Maturity_Picked_raw"].astype(str).str.strip() == ""
    )
    df_all.loc[missing_maturity, "Maturity_Picked_raw"] = df_all.loc[
        missing_maturity, "Original_Line_Maturity_Date__c"
    ]
    df_all["Maturity_dt"] = df_all["Maturity_Picked_raw"].apply(parse_date_any)

    df_term_raw, df_bridge_opp = classify_term_bridge(df_all)

    if not df_term_raw.empty:
        term_ids = df_term_raw["Id"].dropna().astype(str).unique().tolist()
        df_contacts = query_deal_contacts_for_guarantors(sf, term_ids)
        if not df_contacts.empty:
            guarantor_map = (
                df_contacts.groupby(DC_DEAL_FIELD)["GuarantorName"]
                .apply(
                    lambda s: ", ".join(
                        pd.unique([item for item in s.tolist() if str(item).strip() != ""])
                    )
                )
                .reset_index()
                .rename(columns={DC_DEAL_FIELD: "Id", "GuarantorName": "Guarantor"})
            )
            df_term_raw = df_term_raw.merge(guarantor_map, on="Id", how="left")
        else:
            df_term_raw["Guarantor"] = ""
    else:
        df_term_raw["Guarantor"] = ""

    df_term = pd.DataFrame()
    if not df_term_raw.empty:
        df_term["Loan ID"] = df_term_raw["Deal_Loan_Number__c"].apply(
            lambda x: str(last5_strip_prefix(x)).zfill(5)
            if str(last5_strip_prefix(x)).strip()
            else ""
        )
        df_term["Loan"] = df_term_raw.get("Name", "")
        df_term["Account Name"] = df_term_raw.get("Account_Name__c", "")
        df_term["Guarantor"] = df_term_raw.get("Guarantor", "").fillna("")
        df_term["Origination Date"] = df_term_raw.get("OriginationDate_dt")
        df_term["Maturity Date"] = df_term_raw.get("Maturity_dt")
        df_term["Next Payment Date"] = df_term_raw.get("NextPay_dt")
        df_term["Loan Amount Num"] = pd.to_numeric(df_term_raw.get("Amount"), errors="coerce")
        df_term["Outstanding Balance Num"] = pd.to_numeric(
            df_term_raw.get("Current_UPB__c"), errors="coerce"
        )
        df_term["LTV Dec"] = df_term_raw.get("LTV_Dec")
        df_term["Rate Dec"] = df_term_raw.get("Rate_Dec")
        df_term["State(s)"] = df_term_raw.get("State_Percentages__c").apply(extract_states_only)
        df_term["Total Properties Num"] = (
            pd.to_numeric(df_term_raw.get("Total_Properties__c"), errors="coerce")
            .round(0)
            .astype("Int64")
        )
        df_term["Total Units Num"] = (
            pd.to_numeric(df_term_raw.get("Total_Units__c"), errors="coerce")
            .round(0)
            .astype("Int64")
        )
        df_term["Recourse"] = df_term_raw.get("Recourse__c")
        df_term["Historical Ontime % Dec"] = df_term_raw.get(
            "Historical_Ontime_Payments_Percentage__c"
        ).apply(pct_to_dec)
        df_term = df_term.sort_values(
            ["Origination Date", "Loan ID"],
            ascending=[False, True],
            kind="stable",
        ).reset_index(drop=True)

    df_bridge = pd.DataFrame()
    if not df_bridge_opp.empty:
        deal_ids = df_bridge_opp["Id"].dropna().astype(str).unique().tolist()
        advance_fields = ["Id", "Deal__c", "Advance_Num__c", "LOC_Commitment__c", "Wire_Date__c"]
        advance_rows: list[dict[str, Any]] = []
        for group in chunked(deal_ids, 200):
            where_advance = f"Deal__c IN ({', '.join(soql_quote(item) for item in group)})"
            rows_adv, _, _ = try_query_drop_missing(
                sf,
                "Advance__c",
                advance_fields,
                where_advance,
                limit=2000,
                order_by="CreatedDate DESC",
            )
            advance_rows.extend(rows_adv)
        df_adv = pd.DataFrame(advance_rows).drop(columns=["attributes"], errors="ignore")

        if df_adv.empty:
            df_bridge["Loan ID"] = df_bridge_opp["Deal_Loan_Number__c"].apply(
                lambda x: str(last5_strip_prefix(x)).zfill(5)
                if str(last5_strip_prefix(x)).strip()
                else ""
            )
            df_bridge["Loan"] = df_bridge_opp.get("Name", "")
            df_bridge["Account Name"] = df_bridge_opp.get("Account_Name__c", "")
            df_bridge["Commitment Amount Num"] = None
            df_bridge["Origination Date"] = df_bridge_opp.get("OriginationDate_dt")
            df_bridge["Maturity Date"] = df_bridge_opp.get("Maturity_dt")
            df_bridge["Rate Dec"] = df_bridge_opp.get("Rate_Dec")
            df_bridge["LTV Dec"] = df_bridge_opp.get("LTV_Dec")
            df_bridge["Advances Num"] = None
            df_bridge["Total Properties Num"] = None
            df_bridge["State(s)"] = df_bridge_opp.get("State_Percentages__c").apply(extract_states_only)
            df_bridge["Lifetime Funded Num"] = pd.to_numeric(
                df_bridge_opp.get("Aggregate_Funding__c"), errors="coerce"
            )
            df_bridge["Paid Off Assets Num"] = None
            df_bridge["Active Assets Num"] = None
            df_bridge["Outstanding Balance Num"] = pd.to_numeric(
                df_bridge_opp.get("Current_UPB__c"), errors="coerce"
            )
            df_bridge["As-Is/ ARV Num"] = None
            df_bridge["Avg Hold Time Num"] = None
            df_bridge["Avg Disposed Time Num"] = None
        else:
            from datetime import date

            df_adv["Advance_Num__c"] = pd.to_numeric(df_adv.get("Advance_Num__c"), errors="coerce")
            df_adv["LOC_Commitment__c"] = pd.to_numeric(
                df_adv.get("LOC_Commitment__c"), errors="coerce"
            )
            df_adv["Wire_Date__c_dt"] = pd.to_datetime(df_adv.get("Wire_Date__c"), errors="coerce")

            advance_roll = (
                df_adv.groupby("Deal__c", dropna=False)
                .agg(Commitment=("LOC_Commitment__c", "max"), Advances=("Advance_Num__c", "max"))
                .reset_index()
            )

            property_fields = [
                "Id",
                "Advance__c",
                "Deal__c",
                "Payoff_Received_Date__c",
                "After_Repair_Value__c",
                "Appraised_Value_Amount__c",
            ]
            property_rows: list[dict[str, Any]] = []
            advance_ids = df_adv["Id"].dropna().astype(str).unique().tolist()
            for group in chunked(advance_ids, 200):
                where_property = f"Advance__c IN ({', '.join(soql_quote(item) for item in group)})"
                rows_prop, _, _ = try_query_drop_missing(
                    sf,
                    "Property__c",
                    property_fields,
                    where_property,
                    limit=2000,
                    order_by="CreatedDate DESC",
                )
                property_rows.extend(rows_prop)
            df_prop = pd.DataFrame(property_rows).drop(columns=["attributes"], errors="ignore")
            today_dt = pd.to_datetime(date.today())

            if df_prop.empty:
                property_metrics = pd.DataFrame(
                    columns=[
                        "Deal__c",
                        "Total_Assets",
                        "Paid_Off",
                        "Active",
                        "AsIs_ARV",
                        "Avg_Hold",
                        "Avg_Disposed",
                    ]
                )
            else:
                df_prop["Payoff_dt"] = pd.to_datetime(
                    df_prop.get("Payoff_Received_Date__c"), errors="coerce"
                )
                df_prop["ARV_num"] = pd.to_numeric(
                    df_prop.get("After_Repair_Value__c"), errors="coerce"
                )
                missing_arv = df_prop["ARV_num"].isna()
                df_prop.loc[missing_arv, "ARV_num"] = pd.to_numeric(
                    df_prop.get("Appraised_Value_Amount__c"), errors="coerce"
                )
                df_prop = df_prop.merge(
                    df_adv[["Id", "Deal__c", "Wire_Date__c_dt"]],
                    left_on="Advance__c",
                    right_on="Id",
                    how="left",
                    suffixes=("", "_adv"),
                )
                property_id_col = "Id_x" if "Id_x" in df_prop.columns else "Id"

                earliest_wire = (
                    df_prop.groupby(["Deal__c", property_id_col], dropna=False)["Wire_Date__c_dt"]
                    .min()
                    .reset_index()
                    .rename(columns={property_id_col: "PropertyId", "Wire_Date__c_dt": "EarliestWire"})
                )
                payoff_per_asset = (
                    df_prop.groupby(["Deal__c", property_id_col], dropna=False)["Payoff_dt"]
                    .min()
                    .reset_index()
                    .rename(columns={property_id_col: "PropertyId", "Payoff_dt": "Payoff"})
                )
                arv_per_asset = (
                    df_prop.groupby(["Deal__c", property_id_col], dropna=False)["ARV_num"]
                    .max()
                    .reset_index()
                    .rename(columns={property_id_col: "PropertyId", "ARV_num": "ARV"})
                )
                assets = earliest_wire.merge(
                    payoff_per_asset,
                    on=["Deal__c", "PropertyId"],
                    how="left",
                ).merge(
                    arv_per_asset,
                    on=["Deal__c", "PropertyId"],
                    how="left",
                )
                assets["Is_Active"] = assets["Payoff"].isna()
                assets["Is_PaidOff"] = assets["Payoff"].notna()
                assets["Hold_Days"] = (today_dt - assets["EarliestWire"]).dt.days
                assets.loc[assets["EarliestWire"].isna(), "Hold_Days"] = pd.NA
                assets["Disposed_Days"] = (assets["Payoff"] - assets["EarliestWire"]).dt.days
                assets.loc[
                    assets["Payoff"].isna() | assets["EarliestWire"].isna(),
                    "Disposed_Days",
                ] = pd.NA
                base = (
                    assets.groupby("Deal__c", dropna=False)
                    .agg(
                        Total_Assets=("PropertyId", "nunique"),
                        Paid_Off=("Is_PaidOff", "sum"),
                        Active=("Is_Active", "sum"),
                        AsIs_ARV=("ARV", "sum"),
                    )
                    .reset_index()
                )
                hold = (
                    assets[assets["Is_Active"]]
                    .groupby("Deal__c", dropna=False)
                    .agg(Avg_Hold=("Hold_Days", "mean"))
                    .reset_index()
                )
                disposed = (
                    assets[assets["Is_PaidOff"]]
                    .groupby("Deal__c", dropna=False)
                    .agg(Avg_Disposed=("Disposed_Days", "mean"))
                    .reset_index()
                )
                property_metrics = base.merge(hold, on="Deal__c", how="left").merge(
                    disposed,
                    on="Deal__c",
                    how="left",
                )

            bridge_base = df_bridge_opp.copy()
            bridge_base = bridge_base.merge(
                advance_roll,
                left_on="Id",
                right_on="Deal__c",
                how="left",
            ).drop(columns=["Deal__c"], errors="ignore")
            bridge_base = bridge_base.merge(
                property_metrics,
                left_on="Id",
                right_on="Deal__c",
                how="left",
            ).drop(columns=["Deal__c"], errors="ignore")

            df_bridge["Loan ID"] = bridge_base["Deal_Loan_Number__c"].apply(
                lambda x: str(last5_strip_prefix(x)).zfill(5)
                if str(last5_strip_prefix(x)).strip()
                else ""
            )
            df_bridge["Loan"] = bridge_base.get("Name", "")
            df_bridge["Account Name"] = bridge_base.get("Account_Name__c", "")
            df_bridge["Commitment Amount Num"] = pd.to_numeric(
                bridge_base.get("Commitment"), errors="coerce"
            )
            df_bridge["Origination Date"] = bridge_base.get("OriginationDate_dt")
            df_bridge["Maturity Date"] = bridge_base.get("Maturity_dt")
            df_bridge["Rate Dec"] = bridge_base.get("Rate_Dec")
            df_bridge["LTV Dec"] = bridge_base.get("LTV_Dec")
            df_bridge["Advances Num"] = (
                pd.to_numeric(bridge_base.get("Advances"), errors="coerce").round(0).astype("Int64")
            )
            df_bridge["Total Properties Num"] = (
                pd.to_numeric(bridge_base.get("Total_Assets"), errors="coerce")
                .round(0)
                .astype("Int64")
            )
            df_bridge["State(s)"] = bridge_base.get("State_Percentages__c").apply(extract_states_only)
            df_bridge["Lifetime Funded Num"] = pd.to_numeric(
                bridge_base.get("Aggregate_Funding__c"), errors="coerce"
            )
            df_bridge["Paid Off Assets Num"] = (
                pd.to_numeric(bridge_base.get("Paid_Off"), errors="coerce").round(0).astype("Int64")
            )
            df_bridge["Active Assets Num"] = (
                pd.to_numeric(bridge_base.get("Active"), errors="coerce").round(0).astype("Int64")
            )
            df_bridge["Outstanding Balance Num"] = pd.to_numeric(
                bridge_base.get("Current_UPB__c"), errors="coerce"
            )
            df_bridge["As-Is/ ARV Num"] = pd.to_numeric(
                bridge_base.get("AsIs_ARV"), errors="coerce"
            )
            df_bridge["Avg Hold Time Num"] = (
                pd.to_numeric(bridge_base.get("Avg_Hold"), errors="coerce").round(0).astype("Int64")
            )
            df_bridge["Avg Disposed Time Num"] = (
                pd.to_numeric(bridge_base.get("Avg_Disposed"), errors="coerce")
                .round(0)
                .astype("Int64")
            )

        df_bridge = df_bridge.sort_values(
            ["Origination Date", "Loan ID"],
            ascending=[False, True],
            kind="stable",
        ).reset_index(drop=True)

    return df_term, df_bridge


# -------------------------
# Occupancy helpers
# -------------------------
def quarter_label(period_end_date: pd.Timestamp) -> str:
    quarter = ((period_end_date.month - 1) // 3) + 1
    return f"{period_end_date.year} Q{quarter}"



def build_occupancy_lookup(
    berkadia_bytes: bytes,
    periods_to_keep: int = 4,
    min_coverage_ratio: float = 0.25,
):
    required_columns = [
        "Investor Loan#",
        "Consolidated?",
        "Prop Seq#",
        "Property Name",
        "Freq of Analysis",
        "Period End Date",
        "Occupancy %",
    ]
    df = pd.read_excel(
        io.BytesIO(berkadia_bytes),
        sheet_name="Financial Analysis",
        header=3,
        usecols=lambda c: c in required_columns,
    )

    missing_columns = [column for column in required_columns if column not in df.columns]
    if missing_columns:
        raise ValueError(
            "The Financial Analysis sheet is missing required columns: " + ", ".join(missing_columns)
        )

    df["Loan ID"] = df["Investor Loan#"].apply(
        lambda x: str(last5_strip_prefix(x)).zfill(5) if str(last5_strip_prefix(x)).strip() else None
    )
    df["Period End Date"] = pd.to_datetime(df["Period End Date"], errors="coerce")
    df["Occupancy Dec"] = df["Occupancy %"].apply(pct_to_dec)
    df["Prop Seq#"] = pd.to_numeric(df["Prop Seq#"], errors="coerce")
    df["Is Consolidated"] = (
        df["Consolidated?"].fillna("").astype(str).str.strip().str.upper().eq("Y")
    )
    df = df.dropna(subset=["Loan ID", "Period End Date"]).copy()
    if df.empty:
        raise ValueError("No usable occupancy rows were found in the Financial Analysis sheet.")

    def choose_row(group: pd.DataFrame) -> pd.Series:
        consolidated_rows = group[group["Is Consolidated"]]
        if not consolidated_rows.empty:
            row = consolidated_rows.sort_values(["Prop Seq#", "Property Name"], na_position="last").iloc[0].copy()
            row["Occupancy Source"] = "Consolidated row"
            return row

        non_null_occ = group[group["Occupancy Dec"].notna()].copy()
        if non_null_occ.empty:
            row = group.sort_values(["Prop Seq#", "Property Name"], na_position="last").iloc[0].copy()
            row["Occupancy Source"] = "No occupancy value"
            return row

        if len(non_null_occ) == 1:
            row = non_null_occ.iloc[0].copy()
            row["Occupancy Source"] = "Single property row"
            return row

        row = non_null_occ.sort_values(["Prop Seq#", "Property Name"], na_position="last").iloc[0].copy()
        row["Occupancy Dec"] = non_null_occ["Occupancy Dec"].mean()
        row["Occupancy Source"] = "Average of property rows"
        return row

    reduced_rows = []
    for _, group in df.groupby(["Loan ID", "Period End Date"], dropna=False):
        reduced_rows.append(choose_row(group))
    reduced = pd.DataFrame(reduced_rows)

    reduced["Period Label"] = reduced["Period End Date"].apply(quarter_label)
    period_summary = (
        reduced.groupby(["Period Label", "Period End Date"], dropna=False)
        .agg(
            Loan_Count=("Loan ID", "nunique"),
            Occupancy_Count=("Occupancy Dec", lambda s: int(s.notna().sum())),
            Frequency_Types=(
                "Freq of Analysis",
                lambda s: ", ".join(sorted({str(v) for v in s if pd.notna(v) and str(v).strip() != ""})),
            ),
        )
        .reset_index()
        .sort_values(["Period End Date", "Period Label"], ascending=[False, False])
        .reset_index(drop=True)
    )

    max_coverage = int(period_summary["Occupancy_Count"].max()) if not period_summary.empty else 0
    minimum_coverage = max(1, int(round(max_coverage * min_coverage_ratio)))

    selected_summary = period_summary[period_summary["Occupancy_Count"] >= minimum_coverage].copy()
    if len(selected_summary) < periods_to_keep:
        selected_summary = period_summary.copy()
    selected_summary = selected_summary.head(periods_to_keep).copy()

    period_labels = selected_summary["Period Label"].tolist()
    recent_periods = [pd.Timestamp(period) for period in selected_summary["Period End Date"].tolist()]

    reduced = reduced[reduced["Period End Date"].isin(recent_periods)].copy()
    pivot = (
        reduced.pivot_table(
            index="Loan ID",
            columns="Period Label",
            values="Occupancy Dec",
            aggfunc="first",
        )
        .reindex(columns=period_labels)
        .reset_index()
    )

    debug_columns = [
        "Loan ID",
        "Period End Date",
        "Period Label",
        "Occupancy %",
        "Occupancy Dec",
        "Occupancy Source",
        "Consolidated?",
        "Prop Seq#",
        "Property Name",
        "Freq of Analysis",
    ]
    debug_df = reduced[debug_columns].sort_values(
        ["Loan ID", "Period End Date"],
        ascending=[True, False],
    )

    period_summary["Selected"] = period_summary["Period Label"].isin(period_labels)
    period_summary["Coverage Threshold"] = minimum_coverage
    return pivot, period_labels, debug_df, period_summary


@st.cache_data(show_spinner=False)
def load_occupancy_lookup_cached(berkadia_bytes: bytes):
    return build_occupancy_lookup(berkadia_bytes)



def add_occupancy_to_term_rows(
    term_rows: pd.DataFrame,
    occupancy_pivot: pd.DataFrame,
    period_labels: list[str],
) -> pd.DataFrame:
    if term_rows.empty:
        return term_rows.copy()

    result = term_rows.copy()
    lookup = occupancy_pivot.copy()
    for label in period_labels:
        if label not in lookup.columns:
            lookup[label] = None

    lookup = lookup.set_index("Loan ID")
    for label in period_labels:
        result[f"{label} Occ%"] = result["Loan ID"].map(lookup[label])

    occ_cols = [f"{label} Occ%" for label in period_labels]
    result["Occupancy Matched"] = result[occ_cols].notna().any(axis=1)
    return result


# -------------------------
# Excel helpers
# -------------------------
def norm_hdr(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return re.sub(r"\s+", " ", text)



def find_header_row_and_map(ws, must_have=("portfolio", "loan id"), scan_rows: int = 80, scan_cols: int = 200):
    header_row = None
    for row_num in range(1, min(ws.max_row, scan_rows) + 1):
        row_values = [norm_hdr(ws.cell(row_num, col_num).value) for col_num in range(1, min(ws.max_column, scan_cols) + 1)]
        if all(header in row_values for header in must_have):
            header_row = row_num
            break
    if header_row is None:
        raise ValueError(f"Could not find header row on '{ws.title}' containing {must_have}.")

    col_map = {}
    for col_num in range(1, min(ws.max_column, scan_cols) + 1):
        value = ws.cell(header_row, col_num).value
        if value is not None and str(value).strip() != "":
            col_map[norm_hdr(value)] = col_num
    return header_row, col_map



def find_total_row(ws, header_row: int, scan_cols: int = 200):
    start_row = header_row + 1
    for row_num in range(start_row, ws.max_row + 1):
        for col_num in range(1, min(ws.max_column, scan_cols) + 1):
            value = ws.cell(row_num, col_num).value
            if isinstance(value, str) and value.strip().lower() == "total":
                return row_num
    return None



def snapshot_row_style(ws, row_num: int, last_col: int):
    row_height = ws.row_dimensions[row_num].height
    styles = {}
    for col_num in range(1, last_col + 1):
        cell = ws.cell(row_num, col_num)
        styles[col_num] = {
            "_style": copy(cell._style),
            "font": copy(cell.font),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
            "number_format": cell.number_format,
        }
    return row_height, styles



def apply_row_style(ws, row_num: int, styles_by_col, row_height, last_col: int) -> None:
    if row_height is not None:
        ws.row_dimensions[row_num].height = row_height
    for col_num in range(1, last_col + 1):
        cell = ws.cell(row_num, col_num)
        style = styles_by_col.get(col_num)
        if not style:
            continue
        cell._style = copy(style["_style"])
        cell.font = copy(style["font"])
        cell.border = copy(style["border"])
        cell.fill = copy(style["fill"])
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
        cell.number_format = style["number_format"]



def ensure_rows(ws, header_row: int, total_row: int, needed_rows: int, last_col: int):
    if total_row is None:
        raise ValueError(f"Could not find TOTAL row on '{ws.title}'.")

    start_row = header_row + 1
    existing_rows = total_row - start_row
    row_a = start_row
    row_b = start_row + 1 if start_row + 1 < total_row else start_row
    a_height, a_styles = snapshot_row_style(ws, row_a, last_col)
    b_height, b_styles = snapshot_row_style(ws, row_b, last_col)

    if needed_rows > existing_rows:
        add_rows = needed_rows - existing_rows
        ws.insert_rows(total_row, amount=add_rows)
        total_row += add_rows
    elif needed_rows < existing_rows:
        remove_rows = existing_rows - needed_rows
        ws.delete_rows(total_row - remove_rows, amount=remove_rows)
        total_row -= remove_rows

    for idx, row_num in enumerate(range(start_row, start_row + needed_rows)):
        use_alternate = idx % 2 == 1
        apply_row_style(
            ws,
            row_num,
            b_styles if use_alternate else a_styles,
            b_height if use_alternate else a_height,
            last_col,
        )

    return start_row, total_row



def excel_safe(value: Any):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        import numpy as np

        if isinstance(value, np.generic):
            return value.item()
    except Exception:
        pass
    return value



def set_cell(ws, row_num: int, col_num: int, value: Any, number_format: str | None = None) -> None:
    cell = ws.cell(row_num, col_num)
    cell.value = excel_safe(value)
    if number_format:
        cell.number_format = number_format



def sum_ints(series) -> int:
    values = [value for value in series if value is not None and not pd.isna(value)]
    return int(sum(values)) if values else 0



def sum_money(series) -> float:
    values = [float(value) for value in series if value is not None and not pd.isna(value)]
    return float(sum(values)) if values else 0.0



def set_term_occupancy_headers(ws, header_row: int, col_map: dict[str, int], period_labels: list[str]) -> list[int]:
    occ_cols = [col_num for header, col_num in col_map.items() if "occ" in header]
    occ_cols = sorted(occ_cols)
    for index, col_num in enumerate(occ_cols):
        if index < len(period_labels):
            ws.cell(header_row, col_num).value = f"{period_labels[index]} Occ%"
    return occ_cols



def write_term_sheet(ws, term_rows: pd.DataFrame, period_labels: list[str]):
    header_row, col_map = find_header_row_and_map(ws, must_have=("portfolio", "loan id"))
    occ_cols = set_term_occupancy_headers(ws, header_row, col_map, period_labels)
    last_col = max(max(col_map.values()), max(occ_cols) if occ_cols else 0)
    total_row = find_total_row(ws, header_row)
    start_row, total_row = ensure_rows(ws, header_row, total_row, needed_rows=len(term_rows), last_col=last_col)

    def col(name: str):
        return col_map.get(norm_hdr(name))

    total_loan_amount = sum_money(term_rows["Loan Amount Num"].tolist()) if "Loan Amount Num" in term_rows else 0.0
    total_upb = sum_money(term_rows["Outstanding Balance Num"].tolist()) if "Outstanding Balance Num" in term_rows else 0.0
    total_properties = sum_ints(term_rows["Total Properties Num"].tolist()) if "Total Properties Num" in term_rows else 0
    total_units = sum_ints(term_rows["Total Units Num"].tolist()) if "Total Units Num" in term_rows else 0

    for idx, row in term_rows.reset_index(drop=True).iterrows():
        row_num = start_row + idx
        if col("portfolio"):
            set_cell(ws, row_num, col("portfolio"), "Term")
        if col("loan id"):
            set_cell(ws, row_num, col("loan id"), row.get("Loan ID", ""))
        if col("loan"):
            set_cell(ws, row_num, col("loan"), row.get("Loan", ""))
        if col("account name"):
            set_cell(ws, row_num, col("account name"), row.get("Account Name", ""))
        if col("guarantor"):
            set_cell(ws, row_num, col("guarantor"), row.get("Guarantor", ""))
        if col("origination date"):
            set_cell(ws, row_num, col("origination date"), row.get("Origination Date", None), "m/d/yyyy")
        if col("loan amount"):
            value = row.get("Loan Amount Num", None)
            value = None if (value is None or pd.isna(value)) else int(round(float(value)))
            set_cell(ws, row_num, col("loan amount"), value, "$#,##0")
        if col("outstanding balance"):
            value = row.get("Outstanding Balance Num", None)
            value = None if (value is None or pd.isna(value)) else int(round(float(value)))
            set_cell(ws, row_num, col("outstanding balance"), value, "$#,##0")
        if col("origination ltv"):
            value = row.get("LTV Dec", None)
            value = None if (value is None or pd.isna(value)) else float(value)
            set_cell(ws, row_num, col("origination ltv"), value, "0%")
        if col("interest rate"):
            value = row.get("Rate Dec", None)
            value = None if (value is None or pd.isna(value)) else float(value)
            set_cell(ws, row_num, col("interest rate"), value, "0.00%")
        if col("state(s)"):
            set_cell(ws, row_num, col("state(s)"), row.get("State(s)", ""))
        if col("total properties"):
            set_cell(ws, row_num, col("total properties"), row.get("Total Properties Num", None), "0")
        if col("total units"):
            set_cell(ws, row_num, col("total units"), row.get("Total Units Num", None), "0")
        if col("recourse"):
            set_cell(ws, row_num, col("recourse"), row.get("Recourse", ""))
        if col("historical ontime payment %"):
            value = row.get("Historical Ontime % Dec", None)
            value = None if (value is None or pd.isna(value)) else float(value)
            set_cell(ws, row_num, col("historical ontime payment %"), value, "0%")
        if col("next payment date"):
            outstanding_balance = row.get("Outstanding Balance Num", None)
            is_paid_off = False
            if outstanding_balance is not None and not pd.isna(outstanding_balance):
                try:
                    is_paid_off = float(outstanding_balance) == 0.0
                except Exception:
                    is_paid_off = False
            if is_paid_off:
                set_cell(ws, row_num, col("next payment date"), "Paid Off")
            else:
                set_cell(ws, row_num, col("next payment date"), row.get("Next Payment Date", None), "m/d/yyyy")
        if col("current loan maturity date"):
            set_cell(ws, row_num, col("current loan maturity date"), row.get("Maturity Date", None), "m/d/yyyy")

        for occ_index, label in enumerate(period_labels):
            if occ_index >= len(occ_cols):
                break
            value = row.get(f"{label} Occ%", None)
            value = None if (value is None or pd.isna(value)) else float(value)
            set_cell(ws, row_num, occ_cols[occ_index], value, "0.0%")

        for extra_col in occ_cols[len(period_labels) :]:
            set_cell(ws, row_num, extra_col, None)

    if col("loan"):
        set_cell(ws, total_row, col("loan"), int(len(term_rows)))
    if col("loan amount"):
        set_cell(ws, total_row, col("loan amount"), int(round(total_loan_amount)) if total_loan_amount else 0, "$#,##0")
    if col("outstanding balance"):
        set_cell(ws, total_row, col("outstanding balance"), int(round(total_upb)) if total_upb else 0, "$#,##0")
    if col("total properties"):
        set_cell(ws, total_row, col("total properties"), total_properties, "0")
    if col("total units"):
        set_cell(ws, total_row, col("total units"), total_units, "0")



def write_bridge_sheet(ws, bridge_rows: pd.DataFrame):
    header_row, col_map = find_header_row_and_map(ws, must_have=("portfolio", "loan id"))
    last_col = max(col_map.values())
    total_row = find_total_row(ws, header_row)
    start_row, total_row = ensure_rows(ws, header_row, total_row, needed_rows=len(bridge_rows), last_col=last_col)

    def col(name: str):
        return col_map.get(norm_hdr(name))

    total_commitment = sum_money(bridge_rows["Commitment Amount Num"].tolist()) if "Commitment Amount Num" in bridge_rows else 0.0
    total_lifetime = sum_money(bridge_rows["Lifetime Funded Num"].tolist()) if "Lifetime Funded Num" in bridge_rows else 0.0
    total_upb = sum_money(bridge_rows["Outstanding Balance Num"].tolist()) if "Outstanding Balance Num" in bridge_rows else 0.0
    total_properties = sum_ints(bridge_rows["Total Properties Num"].tolist()) if "Total Properties Num" in bridge_rows else 0
    total_paid = sum_ints(bridge_rows["Paid Off Assets Num"].tolist()) if "Paid Off Assets Num" in bridge_rows else 0
    total_active = sum_ints(bridge_rows["Active Assets Num"].tolist()) if "Active Assets Num" in bridge_rows else 0

    for idx, row in bridge_rows.reset_index(drop=True).iterrows():
        row_num = start_row + idx
        if col("portfolio"):
            set_cell(ws, row_num, col("portfolio"), "Bridge")
        if col("loan id"):
            set_cell(ws, row_num, col("loan id"), row.get("Loan ID", ""))
        if col("loan name"):
            set_cell(ws, row_num, col("loan name"), row.get("Loan", ""))
        elif col("loan"):
            set_cell(ws, row_num, col("loan"), row.get("Loan", ""))
        if col("commitment amount"):
            value = row.get("Commitment Amount Num", None)
            value = None if (value is None or pd.isna(value)) else int(round(float(value)))
            set_cell(ws, row_num, col("commitment amount"), value, "$#,##0")
        if col("line origination date"):
            set_cell(ws, row_num, col("line origination date"), row.get("Origination Date", None), "m/d/yyyy")
        elif col("origination date"):
            set_cell(ws, row_num, col("origination date"), row.get("Origination Date", None), "m/d/yyyy")
        if col("line maturity date"):
            set_cell(ws, row_num, col("line maturity date"), row.get("Maturity Date", None), "m/d/yyyy")
        if col("interest rate"):
            value = row.get("Rate Dec", None)
            value = None if (value is None or pd.isna(value)) else float(value)
            set_cell(ws, row_num, col("interest rate"), value, "0.00%")
        if col("ltv"):
            value = row.get("LTV Dec", None)
            value = None if (value is None or pd.isna(value)) else float(value)
            set_cell(ws, row_num, col("ltv"), value, "0%")
        if col("advances"):
            set_cell(ws, row_num, col("advances"), row.get("Advances Num", None), "0")
        if col("total funded assets"):
            set_cell(ws, row_num, col("total funded assets"), row.get("Total Properties Num", None), "0")
        if col("state(s)"):
            set_cell(ws, row_num, col("state(s)"), row.get("State(s)", ""))
        if col("lifetime funded"):
            value = row.get("Lifetime Funded Num", None)
            value = None if (value is None or pd.isna(value)) else int(round(float(value)))
            set_cell(ws, row_num, col("lifetime funded"), value, "$#,##0")
        if col("paid off assets"):
            set_cell(ws, row_num, col("paid off assets"), row.get("Paid Off Assets Num", None), "0")
        if col("active assets"):
            set_cell(ws, row_num, col("active assets"), row.get("Active Assets Num", None), "0")
        if col("outstanding balance"):
            value = row.get("Outstanding Balance Num", None)
            value = None if (value is None or pd.isna(value)) else int(round(float(value)))
            set_cell(ws, row_num, col("outstanding balance"), value, "$#,##0")
        if col("as-is/ arv"):
            value = row.get("As-Is/ ARV Num", None)
            value = None if (value is None or pd.isna(value)) else int(round(float(value)))
            set_cell(ws, row_num, col("as-is/ arv"), value, "$#,##0")
        if col("avg hold time"):
            set_cell(ws, row_num, col("avg hold time"), row.get("Avg Hold Time Num", None), "0")
        if col("avg disposed time"):
            set_cell(ws, row_num, col("avg disposed time"), row.get("Avg Disposed Time Num", None), "0")

    if col("loan name"):
        set_cell(ws, total_row, col("loan name"), int(len(bridge_rows)))
    elif col("loan"):
        set_cell(ws, total_row, col("loan"), int(len(bridge_rows)))
    if col("commitment amount"):
        set_cell(ws, total_row, col("commitment amount"), int(round(total_commitment)) if total_commitment else 0, "$#,##0")
    if col("lifetime funded"):
        set_cell(ws, total_row, col("lifetime funded"), int(round(total_lifetime)) if total_lifetime else 0, "$#,##0")
    if col("outstanding balance"):
        set_cell(ws, total_row, col("outstanding balance"), int(round(total_upb)) if total_upb else 0, "$#,##0")
    if col("total funded assets"):
        set_cell(ws, total_row, col("total funded assets"), total_properties, "0")
    if col("paid off assets"):
        set_cell(ws, total_row, col("paid off assets"), total_paid, "0")
    if col("active assets"):
        set_cell(ws, total_row, col("active assets"), total_active, "0")



def sanitize_filename(name: str) -> str:
    text = "" if name is None or pd.isna(name) else str(name).strip()
    text = re.sub(r'[<>:"/\\|?*]', "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:120] if text else "AM"



def resolve_repo_template_path() -> Path:
    candidate_paths = [
        Path(__file__).resolve().parent / DEFAULT_TEMPLATE_NAME,
        Path.cwd() / DEFAULT_TEMPLATE_NAME,
        Path(__file__).resolve().parent / "templates" / DEFAULT_TEMPLATE_NAME,
        Path.cwd() / "templates" / DEFAULT_TEMPLATE_NAME,
    ]
    for candidate_path in candidate_paths:
        if candidate_path.exists():
            return candidate_path

    raise RuntimeError(
        "Could not find 'Reference AM Templates.xlsx' in the repository. Place it next to amslide.py or inside a templates folder in the repo."
    )


def load_template_bytes() -> bytes:
    return resolve_repo_template_path().read_bytes()



def build_workbook_bytes(
    template_bytes: bytes,
    term_rows: pd.DataFrame,
    bridge_rows: pd.DataFrame,
    account_name: str,
    period_labels: list[str],
):
    workbook = load_workbook(io.BytesIO(template_bytes))
    if TERM_SHEET in workbook.sheetnames:
        write_term_sheet(workbook[TERM_SHEET], term_rows, period_labels)
    if BRIDGE_SHEET in workbook.sheetnames:
        write_bridge_sheet(workbook[BRIDGE_SHEET], bridge_rows)

    output = io.BytesIO()
    workbook.save(output)
    filename = f"{sanitize_filename(account_name)} AM Slides.xlsx"
    return output.getvalue(), filename



def format_preview(df: pd.DataFrame, period_labels: list[str]) -> pd.DataFrame:
    preview = df.copy()
    for label in period_labels:
        column = f"{label} Occ%"
        if column in preview.columns:
            preview[column] = preview[column].apply(lambda x: "" if pd.isna(x) else f"{x:.1%}")
    if "Historical Ontime % Dec" in preview.columns:
        preview["Historical Ontime % Dec"] = preview["Historical Ontime % Dec"].apply(
            lambda x: "" if pd.isna(x) else f"{x:.0%}"
        )
    if "Occupancy Matched" in preview.columns:
        preview["Occupancy Matched"] = preview["Occupancy Matched"].map({True: "Yes", False: "No"})
    return preview


# -------------------------
# Streamlit UI
# -------------------------
st.set_page_config(page_title="AM Slides Builder", layout="wide")
st.title("AM Slides Builder")
st.caption("Build the AM Slides workbook from Salesforce and a Berkadia Financial Analysis export.")

oauth_config = None
oauth_setup_error = None
try:
    oauth_config = load_salesforce_oauth_config()
    maybe_finish_salesforce_oauth(oauth_config)
except Exception as exc:
    oauth_setup_error = str(exc)

sf = None
if oauth_setup_error is None:
    try:
        sf = get_salesforce_client_from_session()
    except Exception as exc:
        oauth_setup_error = str(exc)

template_error = None
template_path = None
try:
    template_path = resolve_repo_template_path()
except Exception as exc:
    template_error = str(exc)

with st.sidebar:
    st.header("App status")
    if template_error:
        st.error(template_error)
    elif template_path is not None:
        st.success("Template found in repository")
        st.caption(template_path.name)

    st.divider()
    st.header("Salesforce")
    if oauth_setup_error:
        st.error(oauth_setup_error)
    elif sf is None:
        st.info("Not connected")
    else:
        auth_data = st.session_state.get("salesforce_auth", {})
        st.success("Connected")
        if auth_data.get("instance_url"):
            st.caption(auth_data["instance_url"])
        if st.button("Log out of Salesforce", use_container_width=True):
            clear_salesforce_session()
            st.rerun()

if template_error:
    st.error(template_error)
    st.stop()

st.subheader("Step 1: Log in to Salesforce")
if oauth_setup_error:
    st.error(oauth_setup_error)
    st.stop()

if sf is None:
    st.info("Log in first. After you sign in, the app will let you upload the Berkadia file and build the AM slide.")
    st.link_button(
        "Log in to Salesforce",
        build_salesforce_login_url(oauth_config),
        use_container_width=False,
    )
    st.caption(f"Callback URL: {oauth_config['redirect_uri']}")
    st.stop()

st.success("Salesforce login complete.")

st.subheader("Step 2: Upload the Berkadia servicer file")
st.caption("The AM template workbook is loaded automatically from the repository. You do not need to upload it here.")
berkadia_file = st.file_uploader(
    "Upload Berkadia servicer file",
    type=["xlsx", "xlsm"],
    key="berkadia_file",
    help="Use the servicer workbook that contains the Financial Analysis sheet.",
)

if berkadia_file is None:
    st.info("Upload the Berkadia servicer file to continue.")
    st.stop()

occupancy_pivot = None
period_labels = []
occupancy_debug = None
occupancy_period_summary = None
try:
    occupancy_pivot, period_labels, occupancy_debug, occupancy_period_summary = load_occupancy_lookup_cached(
        berkadia_file.getvalue()
    )
    selected_periods = ", ".join(period_labels) if period_labels else "none detected"
    st.success(f"Berkadia file loaded. Using these occupancy periods: {selected_periods}")
except Exception as exc:
    st.error(str(exc))
    st.stop()

if isinstance(occupancy_period_summary, pd.DataFrame) and not occupancy_period_summary.empty:
    with st.expander("Review detected occupancy periods", expanded=False):
        summary_display = occupancy_period_summary.copy()
        summary_display["Selected"] = summary_display["Selected"].map({True: "Yes", False: "No"})
        st.dataframe(summary_display, use_container_width=True, hide_index=True)

st.subheader("Step 3: Search Salesforce and choose an account")
search_col1, search_col2 = st.columns([1, 2])
with search_col1:
    search_mode = st.selectbox(
        "Search Salesforce by",
        ["Account Name", "Deal Name", "Deal Loan Number"],
    )
with search_col2:
    search_text = st.text_input("Search text")

if st.button("Search Salesforce", type="secondary"):
    if not search_text.strip():
        st.error("Enter a search value first.")
    else:
        try:
            account_candidates = search_matching_accounts(sf, search_mode, search_text)
            st.session_state["account_candidates"] = account_candidates
            if account_candidates.empty:
                st.warning("No matching accounts found.")
            else:
                st.success(f"Found {len(account_candidates)} matching account candidates.")
        except Exception as exc:
            st.error(str(normalize_salesforce_error(exc)))

account_candidates = st.session_state.get("account_candidates", pd.DataFrame())
selected_account = None
if isinstance(account_candidates, pd.DataFrame) and not account_candidates.empty:
    st.dataframe(account_candidates, use_container_width=True, hide_index=True)
    selected_account = st.selectbox(
        "Pick the account for the AM slide",
        options=account_candidates["Account_Name__c"].tolist(),
    )
else:
    st.info("Search Salesforce to load the account list.")

st.subheader("Step 4: Build and download the AM slide")
build_disabled = not bool(selected_account)
if st.button("Build completed AM slide", type="primary", disabled=build_disabled):
    try:
        with st.spinner("Building the AM slide workbook..."):
            term_rows, bridge_rows = build_term_bridge_for_account(sf, selected_account)
            if term_rows.empty and bridge_rows.empty:
                raise RuntimeError("No term or bridge rows were returned for the selected account.")

            term_rows_with_occ = add_occupancy_to_term_rows(term_rows, occupancy_pivot, period_labels)
            template_bytes = load_template_bytes()
            workbook_bytes, workbook_name = build_workbook_bytes(
                template_bytes,
                term_rows_with_occ,
                bridge_rows,
                selected_account,
                period_labels,
            )

        st.session_state["term_preview"] = format_preview(term_rows_with_occ, period_labels)
        st.session_state["bridge_preview"] = bridge_rows
        st.session_state["occupancy_debug"] = occupancy_debug
        st.session_state["occupancy_period_summary"] = occupancy_period_summary
        st.session_state["period_labels"] = period_labels
        st.session_state["workbook_bytes"] = workbook_bytes
        st.session_state["workbook_name"] = workbook_name
        st.session_state["match_count"] = int(term_rows_with_occ["Occupancy Matched"].sum()) if not term_rows_with_occ.empty else 0
        st.session_state["term_count"] = len(term_rows_with_occ)

        st.success("The AM slide workbook is ready.")
    except Exception as exc:
        st.error(str(normalize_salesforce_error(exc)))

period_labels = st.session_state.get("period_labels", period_labels)
term_preview = st.session_state.get("term_preview")
bridge_preview = st.session_state.get("bridge_preview")
occupancy_debug = st.session_state.get("occupancy_debug")
occupancy_period_summary = st.session_state.get("occupancy_period_summary", occupancy_period_summary)
workbook_bytes = st.session_state.get("workbook_bytes")
workbook_name = st.session_state.get("workbook_name")

if workbook_bytes:
    metric_col1, metric_col2 = st.columns(2)
    with metric_col1:
        st.metric("Term loans", st.session_state.get("term_count", 0))
    with metric_col2:
        st.metric("Term loans with occupancy match", st.session_state.get("match_count", 0))

    st.download_button(
        "Download completed AM slide (Excel)",
        data=workbook_bytes,
        file_name=workbook_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if isinstance(term_preview, pd.DataFrame) and not term_preview.empty:
    st.subheader("Term preview")
    st.dataframe(term_preview, use_container_width=True, hide_index=True)

if isinstance(bridge_preview, pd.DataFrame) and not bridge_preview.empty:
    st.subheader("Bridge preview")
    st.dataframe(bridge_preview, use_container_width=True, hide_index=True)

if isinstance(occupancy_debug, pd.DataFrame) and not occupancy_debug.empty:
    with st.expander("Occupancy lookup details", expanded=False):
        st.write(
            "Using these period columns from the Berkadia file: " + ", ".join(period_labels)
            if period_labels
            else "No occupancy periods detected."
        )
        debug_display = occupancy_debug.copy()
        debug_display["Occupancy Dec"] = debug_display["Occupancy Dec"].apply(
            lambda x: "" if pd.isna(x) else f"{x:.1%}"
        )
        st.dataframe(debug_display, use_container_width=True, hide_index=True)
